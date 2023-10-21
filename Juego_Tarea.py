
def aleatorio(num1, num2):
    import random as r
    return r.randint(num1, num2)

def escoger_dificultad(): #se escoge al dificultad: el rango de números
    try:
        seleccion = int(input("1. Fácil (0-1000)\n2. Media (0-2000)\n3. Difícil (0-3000)"))
        while seleccion > 3 or seleccion < 1: #valida que se escoja una de las opciones ofrecidas
            seleccion = int(input("Número inválido. Por favor, introduzca un número entre 1 y 3"))
    except:
        seleccion = int(input("Carácter no válido. Por favor, introduzca un número entre 1 y 3"))
        
    #se determina la dificultad según al opción escogida
    if seleccion == 1:
        dificultad = 1000
    elif seleccion == 2:
        dificultad = 2000
    elif seleccion == 3:
        dificultad = 3000
    return dificultad #devuelve el número del rango superior. El inferior siempre será 0.

def escoger_intentos():#selección de número de intentos    
    try:
        seleccion = int(input("1. 20 intentos (fácil)\n2. 12 intentos (medio)\n3. 5 intentos (difícil)"))        
        while seleccion > 3 or seleccion < 1: #valida que se escoja una de las opciones ofrecidas
            seleccion = int(input("Número inválido. Por favor, introduzca un número entre 1 y 3"))
    except:
        seleccion = int(input("Carácter no válido. Por favor, introduzca un número entre 1 y 3"))
    #se determina la dificultad según al opción escogida
    if seleccion == 1:
        intentos = 20
    elif seleccion == 2:
        intentos = 12
    elif seleccion == 3:
        intentos = 5
        
    return intentos #devuelve el número de intentos que tendrá el jugador que adivina

def escoger_jugador():
    jugador1 = str(input('Jugador 1, introduce tu nombre:\n'))
    jugador2 = str(input('Jugador 2, introduce tu nombre:\n'))
    return jugador1, jugador2

def adivinar(intentos, dificultad, solucion):
    queda = intentos
    lista_intentos = [] #registra en una lista los intentos para estadística
    lista_diferencias = []
    for i in range(intentos):
        try:
            intento = int(input('¡Prueba a adivinar el número!'))
        except:
            intento = int(input('Eso no es un número. ¡Prueba de nuevo!'))
        lista_intentos.append(intento) #almacena los intentos en una lista
        diferencia = solucion - intento #diferencia entre el intento y el número solución
        lista_diferencias.append(diferencia) #recoge la diferencia entre el objetivo y el número introducido
                                         #en una lista
        while intento < 0 or intento > dificultad: #comprueba que el número introducido esté en el rango
            print('Número inválido: prueba con un número entre 0 y', dificultad)
            intento = int(input('Prueba de nuevo (no se restan intentos):'))
        if intento > solucion: 
            queda = queda - 1
            print('El número que buscas es menor al que acabas de introducir. Te quedan', queda, 'intentos')
        elif intento < solucion:
            queda = queda - 1
            print('El número que buscas es mayor al que acabas de introducr. Te quedan', queda, 'intentos')
        elif intento == solucion: #si se acierta el número, el resultado es 1. Esto servirá en el futuro para las
            #siguientes funciones
            resultado = 1
            return resultado, lista_intentos, solucion
        
        if queda == 0: #del mismo modo, si no quedan intentos, es que el número no se ha acertado, y el resultado es 2.
            resultado = 2
            
            return resultado, lista_intentos, solucion
        

def modo_solitario(): #juego modo solitario
    import GeneraAleatorio as g
    
    dificultad = escoger_dificultad() #se establece la dificultad
    intentos = escoger_intentos() #los intentos
    solucion = aleatorio(0, dificultad) #se genera aleatoriamente un número aleatorio y este será el número solución
    resultado = adivinar(intentos, dificultad, solucion) 
    
    return resultado 


def dos_jugadores():
    import getpass #esta librería permite esconder el input introducido por el segundo jugador (el número a adivinar)
    dificultad = escoger_dificultad()
    intentos = escoger_intentos()
    try:
        solucion = int(getpass.getpass(prompt = '¡Jugador 2, introduce el número!'))
    except:
        solucion = int(getpass.getpass(prompt = 'Solo se pueden introducir números enteros. ¡Jugador 2, introduce el número!'))
    resultado = adivinar(intentos, dificultad, solucion)
    return resultado

def media_diferencias(lista_intentos, solucion):
    
    #definimos las variables a emplear en esta función
    total = 0
    lista_diferencias = []
    #iteramos por cada objeto en la lista, donde por cada objeto, hacemos la resta (absoluta) con el objetivo
    for i in lista_intentos:
        if i > solucion:
            diferencia = i - solucion
        elif i < solucion:
            diferencia = solucion - i
        lista_diferencias.append(diferencia)
        
    for i in lista_diferencias:
        total = total + i
    
    media_diferencias = total/len(lista_diferencias)
    return media_diferencias

def guardar_datos(resultado, lista_intentos, solucion, modo):
    import openpyxl
    libro = openpyxl.load_workbook('Estadística.xlsx')
    jugador = str(input('Nombre del jugador:\n'))
    
    if jugador not in libro.sheetnames:
        libro.create_sheet(jugador)
        hoja = libro[jugador]
        hoja['A1'].value = 'Modo de juego:'
        hoja['A2'].value = 'Intentos empleados:'
        hoja['A3'].value = 'Desviación media frente al número solución:'
        hoja['A4'].value = 'Número de victorias:'
        hoja['A5'].value = 'Número de derrotas:'
        
        if modo == 1:
            hoja['B1'].value = 'Un jugador'
        elif modo == 2:
            hoja['B1'].value = 'Dos jugadores'
        hoja['B2'].value = 0
        hoja['B3'].value = 0
        hoja['B4'].value = 0
        hoja['B5'].value = 0
    hoja = libro[jugador]
    hoja['B2'].value = len(lista_intentos) #número de intentos empleados. Lista_intentos es la lista de los intentos, números introducidos
    #en la última partida
    hoja['B3'].value = media_diferencias(lista_intentos, solucion) #se emplea la función de calcular la media de las diferencias entre
    #el número introducido y el número objetivo
    
    if resultado == 1:
        hoja['B4'].value = hoja['B4'].value + 1 #suma 1 a las victorias si el jugador 1 ha ganado
        
    elif resultado == 2:
        hoja['B5'].value = hoja['B5'].value + 1 #suma 1 a las derrotas si el jugador 1 ha perdido
    
    libro.save('Estadística.xlsx')
    